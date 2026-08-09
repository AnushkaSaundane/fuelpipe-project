from django.contrib import admin
from django.utils.html import format_html
from .models import Product, ContactMessage, Customer, CompanyImage, Category
from import_export.admin import ImportExportModelAdmin
from .resources import ProductResource
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@admin.register(Product)
class ProductAdmin(ImportExportModelAdmin):
    resource_class = ProductResource

    actions = [
        export_products_with_images
    ]
    list_display = ['name', 'product_type', 'price', 'display_image', 'is_featured','category','part_number']
    list_filter = ['product_type', 'is_featured']
    list_display_links = ['name']
    search_fields = ['part_number','name', 'description']
    list_editable = ('is_featured', 'price')
    list_per_page = 20
    date_hierarchy = 'created_at'
    
    import_template_name = 'admin/import_export/import.html'
    fieldsets = (
        ('Basic Information', {
            'fields': ('part_number', 'product_type','category', 'description', 'image')
        }),
        ('Pricing & Features', {
            'fields': ('price', 'is_featured')
        }),
        ('Specifications', {
            'fields': ('specifications',),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    readonly_fields = ('created_at', 'updated_at')
    
    def display_image(self, obj):
        if obj.image:
            return format_html('<img src="{}" width="50" height="50" style="border-radius: 5px; object-fit: cover;" />', obj.image.url)
        return format_html('<div style="width: 50px; height: 50px; background: #f0f0f0; border-radius: 5px; display: flex; align-items: center; justify-content: center; color: #999;"><i class="fas fa-box"></i></div>')
    display_image.short_description = 'Image'
    
    def status_badge(self, obj):
        if obj.is_featured:
            return format_html('<span style="background: linear-gradient(135deg, #10b981, #059669); color: white; padding: 4px 8px; border-radius: 12px; font-size: 11px; font-weight: 600;">⭐ Featured</span>')
        return format_html('<span style="background: #6b7280; color: white; padding: 4px 8px; border-radius: 12px; font-size: 11px; font-weight: 600;">Standard</span>')
    status_badge.short_description = 'Status'

    def display_image(self, obj):
        if obj.image:
            return format_html('<img src="{}" width="50" height="50" style="border-radius: 5px;" />', obj.image.url)
        return "No Image"
    display_image.short_description = 'Image'

    def export_products_with_images(modeladmin, request, queryset):


        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Products"

        # =====================================================
        # HEADERS
        # =====================================================

        headers = [
            "Part Number",
            "Product Name",
            "Product Type",
            "Category",
            "Description",
            "Specifications",
            "Price",
            "Featured",
            "Image",
        ]

        for col_num, header in enumerate(headers, 1):

            cell = worksheet.cell(
                row=1,
                column=col_num,
                value=header
            )

            cell.font = Font(
                bold=True,
                size=12
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAF7"
            )

        worksheet.row_dimensions[1].height = 30

        # =====================================================
        # COLUMN WIDTHS
        # =====================================================

        worksheet.column_dimensions["A"].width = 20
        worksheet.column_dimensions["B"].width = 30
        worksheet.column_dimensions["C"].width = 20
        worksheet.column_dimensions["D"].width = 20
        worksheet.column_dimensions["E"].width = 45
        worksheet.column_dimensions["F"].width = 45
        worksheet.column_dimensions["G"].width = 15
        worksheet.column_dimensions["H"].width = 12
        worksheet.column_dimensions["I"].width = 25

        # =====================================================
        # PRODUCTS
        # =====================================================

        products = queryset.select_related("category")

        for row_num, product in enumerate(products, start=2):

            # -------------------------------------------------
            # PART NUMBER
            # -------------------------------------------------

            worksheet.cell(
                row=row_num,
                column=1,
                value=product.part_number or ""
            )

            # -------------------------------------------------
            # PRODUCT NAME
            # -------------------------------------------------

            worksheet.cell(
                row=row_num,
                column=2,
                value=product.name or ""
            )

            # -------------------------------------------------
            # PRODUCT TYPE
            # -------------------------------------------------

            product_type = ""

            if product.product_type:
                product_type = product.get_product_type_display()

            worksheet.cell(
                row=row_num,
                column=3,
                value=product_type
            )

            # -------------------------------------------------
            # CATEGORY
            # -------------------------------------------------

            category_name = ""

            if product.category:
                category_name = product.category.name

            worksheet.cell(
                row=row_num,
                column=4,
                value=category_name
            )

            # -------------------------------------------------
            # DESCRIPTION
            # -------------------------------------------------

            worksheet.cell(
                row=row_num,
                column=5,
                value=product.description or ""
            )

            # -------------------------------------------------
            # SPECIFICATIONS
            # -------------------------------------------------

            worksheet.cell(
                row=row_num,
                column=6,
                value=product.specifications or ""
            )

            # -------------------------------------------------
            # PRICE
            # -------------------------------------------------

            price = None

            if product.price is not None:
                price = float(product.price)

            worksheet.cell(
                row=row_num,
                column=7,
                value=price
            )

            # -------------------------------------------------
            # FEATURED
            # -------------------------------------------------

            worksheet.cell(
                row=row_num,
                column=8,
                value="Yes" if product.is_featured else "No"
            )

            # -------------------------------------------------
            # IMAGE
            # -------------------------------------------------

            image_cell = worksheet.cell(
                row=row_num,
                column=9
            )

            image_cell.value = ""

            if product.image:

                try:

                    image_path = product.image.path

                    if os.path.exists(image_path):

                        excel_image = ExcelImage(image_path)

                        # Maximum image size
                        max_width = 150
                        max_height = 100

                        original_width = excel_image.width
                        original_height = excel_image.height

                        # Calculate resize ratio
                        ratio = min(
                            max_width / original_width,
                            max_height / original_height
                        )

                        excel_image.width = int(
                            original_width * ratio
                        )

                        excel_image.height = int(
                            original_height * ratio
                        )

                        # Add actual image to Excel
                        worksheet.add_image(
                            excel_image,
                            f"I{row_num}"
                        )

                        # Increase row height
                        worksheet.row_dimensions[
                            row_num
                        ].height = 85

                except Exception as error:

                    print(
                        f"Could not export image for "
                        f"product {product.id}: {error}"
                    )

            # -------------------------------------------------
            # TEXT ALIGNMENT
            # -------------------------------------------------

            for col_num in range(1, 10):

                cell = worksheet.cell(
                    row=row_num,
                    column=col_num
                )

                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True
                )

        # =====================================================
        # FREEZE HEADER
        # =====================================================

        worksheet.freeze_panes = "A2"

        # =====================================================
        # FILTER
        # =====================================================

        if worksheet.max_row > 1:

            worksheet.auto_filter.ref = (
                f"A1:I{worksheet.max_row}"
            )

        # =====================================================
        # BORDERS
        # =====================================================

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=9
        ):

            for cell in row:
                cell.border = thin_border

        # =====================================================
        # DOWNLOAD
        # =====================================================

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

        response["Content-Disposition"] = (
            'attachment; filename="products_with_images.xlsx"'
        )

        workbook.save(response)

        return response


    export_products_with_images.short_description = (
        "📥 Export Excel with Images"
    )

@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):

    list_display = ['name']



@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    list_display = ['name', 'website', 'is_active']
    list_filter = ['is_active']

@admin.register(CompanyImage)
class CompanyImageAdmin(admin.ModelAdmin):
    list_display = ['name', 'is_active']
    list_filter = ['is_active']

@admin.register(ContactMessage)
class ContactMessageAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'subject', 'created_at', 'is_read']
    list_filter = ['is_read', 'created_at']
    readonly_fields = ['name', 'email', 'phone', 'subject', 'message', 'created_at']

from django.contrib import admin
from .models import ProductRequest

@admin.register(ProductRequest)
class ProductRequestAdmin(admin.ModelAdmin):

    list_display = (
        "part_name",
        "vehicle_company",
        "name",
        "phone",
        "created_at",
    )

    search_fields = (
        "part_name",
        "vehicle_company",
        "name",
        "phone",
    )

    list_filter = ("vehicle_company",)
